import openpyxl as oxl


class Employee:

    def __init__(self):

        dict_data = {
            "Name": ['John', 'Grey', 'Bob', 'France', 'Sam', 'Shirin', 'Shinzu', 'Yamato', 'Kirito', 'Miya'],
            "Emp Id": [65651, 55454, 64656, 46622, 54632, 46252, 66424, 65586, 65848, 64645],
            "Address": ['Washington', 'UK', 'London', 'paris', 'France', 'Colombia', 'Nigiria', 'Akiyabara', 'Tokyo',
                        'Hokaido'],
            "Phone Number": [654632145, 6461666465, 65666464, 6646446, 5462135468, 46634646, 46464, 46462, 65643543,
                             64654354],
            "Email": ['a@gmail.com', 'a@gmail.com', 'afef@gmail.com', 'eafaf@gmail.com', 'esff@gmail.com',
                      'sdfdzf@gmail.com', 'saff@gmail.com', 'afd@gmail.com', 'safdf@gmail.com', 'asfdfd@gmail.com'],

            "Birth of Date": ['15/5/1999', '15/5/1999', '15/5/1999', '15/5/1999', '15/5/1999', '15/5/1999', '15/5/1999',
                              '15/5/1999', '15/5/1999', '15/5/1999'],
            "Designation": ['PAT', 'PAT', 'PAT', 'PAT', 'PAT', 'PAT', 'PAT', 'PAT', 'PAT', 'PAT'],
            "Department": ['testing', 'testing', 'testing', 'testing', 'testing', 'testing', 'testing', 'testing',
                           'testing', 'testing'],
            "Work location": ['Chennai', 'Chennai', 'Chennai', 'Chennai', 'Chennai', 'Chennai', 'Chennai', 'Chennai',
                              'Chennai', 'Chennai'],
            "Salary": [30000, 30000, 30000, 30000, 30000, 30000, 30000, 30000, 30000, 30000],
        }

        self.data = dict_data

    def load_data(self, sheet):
        print('Loading data')
        row_count = 2
        col_count = 1

        for key in self.data.keys():
            sheet.cell(row=1, column=col_count).value = key
            get_key_value = self.data[key]
            for li_data in range(len(get_key_value)):
                sheet.cell(row=row_count, column=col_count).value = get_key_value[li_data]
                row_count += 1
            col_count += 1
            row_count = 2

    def write_data(self):
        first_file = oxl.Workbook()
        first_sheet = first_file.active
        print('Creating File')
        self.load_data(first_sheet)
        print('Saving data')
        first_file.save(filename='FirstSheet.xlsx')
        print('File Created Successfully')

    def read_data(self):
        print('Reading data')
        read_file = oxl.load_workbook('FirstSheet.xlsx')
        read_file_sheet = read_file.active
        read_file_max_row = read_file_sheet.max_row
        read_file_max_col = read_file_sheet.max_column
        print("Max Row : {0}, Max Col : {1}".format(read_file_max_row, read_file_max_col))

        read_data_dict = {}
        col_count = 1
        row_count = 2

        for i in range(read_file_max_row - 1):
            li = []
            get_key = read_file_sheet.cell(row=1, column=col_count).value
            for j in range(read_file_max_row - 1):
                li.append(read_file_sheet.cell(row=row_count, column=col_count).value)
                read_data_dict[get_key] = li
                row_count += 1
            col_count += 1
            row_count = 2

        print(read_data_dict)
        # return read_data_dict

    def create_copy(self):
        print('Creating a copy')
        first_file = oxl.load_workbook('FirstSheet.xlsx')
        first_file_sheet = first_file.worksheets[0]

        second_file = oxl.Workbook()
        second_file_sheet = second_file.active

        max_row = first_file_sheet.max_row
        max_col = first_file_sheet.max_column

        for i in range(1, max_row + 1):
            for j in range(1, max_col + 1):
                second_file_sheet.cell(row=i, column=j).value = first_file_sheet.cell(row=i, column=j).value

        second_file.save(filename='SecondFile.xlsx')
        print('Successfully create a copy of FirstFile.xlsx in SecondFile.xlsx')


if __name__ == "__main__":
    print('Initializing')
    Emp = Employee()
    Emp.write_data()
    print(" ")
    Emp.read_data()
    print(" ")
    Emp.create_copy()
